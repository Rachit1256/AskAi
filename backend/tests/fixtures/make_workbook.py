"""Builds an IMD-shaped workbook used by the ingest tests."""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")


def build(path: Path, years: int = 30) -> Path:
    wb = Workbook()

    # --- Sheet 1: wide monthly series with derived Annual and JJAS columns ---
    ws = wb.active
    ws.title = "Rainfall Series"
    ws["A1"] = "ALL INDIA MONTHLY AND SEASONAL RAINFALL (mm) SERIES"
    ws.merge_cells("A1:P1")
    ws["A1"].font = BOLD
    ws["A3"], ws["B3"] = "Region", "Peninsular India"
    ws["A4"], ws["B4"] = "Units", "mm"

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    header = ["Year", *months, "JJAS", "Annual"]
    for j, name in enumerate(header, start=1):
        cell = ws.cell(row=6, column=j, value=name)
        cell.font, cell.fill = BOLD, HEADER_FILL

    for i in range(years):
        year = 1991 + i
        values = [round(12.0 + (i * 3 + m * 17) % 260 + m * 4.5, 1) for m in range(12)]
        ws.cell(row=7 + i, column=1, value=year)
        for j, v in enumerate(values, start=2):
            ws.cell(row=7 + i, column=j, value=v)
        ws.cell(row=7 + i, column=14, value=round(sum(values[5:9]), 1))  # JJAS
        ws.cell(row=7 + i, column=15, value=round(sum(values), 1))  # Annual

    # --- Sheet 2: station normals, context block + two stacked tables ---
    ws2 = wb.create_sheet("Station Normals")
    ws2["A1"] = "STATION CLIMATOLOGICAL NORMALS"
    ws2.merge_cells("A1:E1")
    ws2["A1"].font = BOLD
    for i, (k, v) in enumerate(
        [
            ("Station", "PUNE"),
            ("Index", 43063),
            ("Period", "1991-2020"),
            ("Prepared", dt.date(2026, 7, 28)),
        ],
        start=3,
    ):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)

    for j, name in enumerate(["Month", "Rainfall (mm)", "Rainy Days", "Max Temp (C)"], start=1):
        cell = ws2.cell(row=9, column=j, value=name)
        cell.font, cell.fill = BOLD, HEADER_FILL
    normals = [
        ("January", 1.9, 0.2, 31.0),
        ("February", 0.4, 0.1, 33.2),
        ("March", 2.1, 0.2, 36.5),
        ("April", 8.9, 0.7, 38.1),
        ("May", 35.2, 2.4, 37.0),
        ("June", 134.7, 8.9, 31.6),
        ("July", 187.3, 14.1, 28.4),
        ("August", 116.5, 12.7, 27.9),
        ("September", 152.8, 8.6, 30.1),
        ("October", 84.6, 4.2, 32.4),
        ("November", 30.1, 1.6, 30.8),
        ("December", 4.3, 0.4, 30.2),
    ]
    for i, row in enumerate(normals, start=10):
        for j, v in enumerate(row, start=1):
            ws2.cell(row=i, column=j, value=v)
    ws2.cell(row=22, column=1, value="Total").font = BOLD
    ws2.cell(row=22, column=2, value=round(sum(r[1] for r in normals), 1))

    # Side-by-side table separated by a blank column
    for j, name in enumerate(["Grade", "Observations"], start=6):
        cell = ws2.cell(row=9, column=j, value=name)
        cell.font, cell.fill = BOLD, HEADER_FILL
    for i, row in enumerate([("A", 310), ("B", 188), ("C", 42)], start=10):
        for j, v in enumerate(row, start=6):
            ws2.cell(row=i, column=j, value=v)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


if __name__ == "__main__":
    print(build(Path("var/sample_imd.xlsx")))
