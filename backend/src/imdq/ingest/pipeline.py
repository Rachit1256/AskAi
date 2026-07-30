"""File -> blocks -> typed long-form records.

Two phases, deliberately separated:

``analyse``  reads a bounded probe window, segments it, types the columns and
             emits an :class:`ExtractionRecipe` per block. Expensive, cached by
             layout fingerprint.
``stream``   replays a recipe over the whole sheet in batches. Cheap, constant
             memory, and never re-infers anything.

Recipes are anchored on header signatures rather than absolute coordinates, so
inserting a comment line above a table does not break extraction.
"""

from __future__ import annotations

import hashlib
from collections.abc import Iterator
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from imdq.domain.imd import MONTH_ABBR
from imdq.errors import IngestFailed, UnsupportedFile
from imdq.ingest.blocks import Block, BlockKind
from imdq.ingest.grid import Grid
from imdq.ingest.normalize import (
    ColumnRole,
    ColumnSpec,
    classify_columns,
    detect_derived_columns,
    parse_number,
    slugify,
)
from imdq.ingest.segmenter import segment

SUPPORTED_SUFFIXES = frozenset({".xlsx", ".xlsm", ".xltx", ".csv", ".tsv"})
BATCH_ROWS = 20_000
ANCHOR_SEARCH_SLACK = 40
SAMPLE_ROWS_FOR_TYPING = 200


def content_hash(path: Path, chunk: int = 1 << 20) -> str:
    """Stream a SHA-256 so re-uploading the same file replaces, never duplicates."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while block := handle.read(chunk):
            digest.update(block)
    return digest.hexdigest()


@dataclass(slots=True)
class ExtractionRecipe:
    sheet: str
    block_id: str
    kind: BlockKind
    header_signature: list[str]
    header_row: int
    body_start: int
    first_col: int
    last_col: int
    columns: list[ColumnSpec]
    context: dict[str, Any] = field(default_factory=dict)
    id_columns: int = 1
    totals_row: int | None = None
    stated_totals: dict[str, float] = field(default_factory=dict)
    confidence: float = 0.0

    @property
    def stored_columns(self) -> list[ColumnSpec]:
        """Derived aggregate columns are excluded from the long form."""
        return [c for c in self.columns if not c.is_derived]

    def to_dict(self) -> dict[str, Any]:
        return {
            "block_id": self.block_id,
            "sheet": self.sheet,
            "kind": str(self.kind),
            "header_row": self.header_row + 1,
            "confidence": round(self.confidence, 3),
            "context": dict(self.context),
            "columns": [c.to_dict() for c in self.columns],
        }


@dataclass(slots=True)
class SheetPlan:
    sheet: str
    fingerprint: str
    recipes: list[ExtractionRecipe]
    blocks: list[Block]
    truncated_probe: bool


@dataclass(slots=True)
class FilePlan:
    path: Path
    filename: str
    content_hash: str
    sheets: list[SheetPlan]

    @property
    def layout_fingerprint(self) -> str:
        joined = "|".join(sheet.fingerprint for sheet in self.sheets)
        return hashlib.sha1(joined.encode()).hexdigest()[:16]

    def to_dict(self) -> dict[str, Any]:
        return {
            "filename": self.filename,
            "content_hash": self.content_hash,
            "layout_fingerprint": self.layout_fingerprint,
            "sheets": [
                {
                    "sheet": s.sheet,
                    "fingerprint": s.fingerprint,
                    "probe_truncated": s.truncated_probe,
                    "blocks": [r.to_dict() for r in s.recipes],
                }
                for s in self.sheets
            ],
        }


@dataclass(slots=True)
class RecordBatch:
    recipe: ExtractionRecipe
    columns: list[str]
    rows: list[tuple[Any, ...]]

    def __len__(self) -> int:
        return len(self.rows)


def sheet_fingerprint(blocks: list[Block]) -> str:
    """Layout hash that ignores row counts, so it survives the data growing."""
    parts = [
        f"{b.kind}|{b.c0}-{b.c1}|{'~'.join(b.header)}"
        for b in sorted(blocks, key=lambda b: (b.r0, b.c0))
    ]
    return hashlib.sha1("\n".join(parts).encode()).hexdigest()[:16]


def _sample_rows(grid: Grid, block: Block) -> list[tuple[Any, ...]]:
    stop = min(block.r1, block.body_start + SAMPLE_ROWS_FOR_TYPING - 1)
    return [
        tuple(grid.values[r, c] for c in range(block.c0, block.c1 + 1))
        for r in range(block.body_start, stop + 1)
        if r != block.totals_row
    ]


def _build_recipe(grid: Grid, block: Block, index: int) -> ExtractionRecipe:
    sample = _sample_rows(grid, block)
    specs = classify_columns(block.header, sample)
    detect_derived_columns(specs, sample)

    for key in block.context:
        if key != "_section":
            specs.append(ColumnSpec(source_name=key, slug=slugify(key), role=ColumnRole.DIMENSION))

    stated_totals: dict[str, float] = {}
    if block.totals_row is not None:
        # The totals row is removed from the data but kept as a free checksum:
        # if it disagrees with our own sum, a sub-table was mis-detected.
        for offset, spec in enumerate(specs[: block.width]):
            parsed = parse_number(grid.values[block.totals_row, block.c0 + offset])
            if isinstance(parsed.value, float):
                stated_totals[spec.slug] = parsed.value

    return ExtractionRecipe(
        sheet=grid.sheet_name,
        block_id=f"{slugify(grid.sheet_name)}_{index:02d}",
        kind=block.kind,
        header_signature=[h for h in block.header if h],
        header_row=block.header_rows[-1] if block.header_rows else block.r0,
        body_start=block.body_start,
        first_col=block.c0,
        last_col=block.c1,
        columns=specs,
        context=dict(block.context),
        totals_row=block.totals_row,
        stated_totals=stated_totals,
        confidence=block.confidence,
    )


def analyse(path: Path, probe_rows: int = 2_000) -> FilePlan:
    """Discover every block in every sheet and emit extraction recipes."""
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise UnsupportedFile(
            f"Unsupported file type '{path.suffix}'.",
            remedy=f"Supported types: {', '.join(sorted(SUPPORTED_SUFFIXES))}",
        )

    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:  # openpyxl raises a wide range of parse errors
        raise IngestFailed(
            f"Could not open workbook: {exc}",
            remedy="Confirm the file is a valid .xlsx and is not password protected.",
        ) from exc

    sheets: list[SheetPlan] = []
    try:
        for worksheet in workbook.worksheets:
            grid = Grid.from_worksheet(worksheet, probe_rows=probe_rows)
            blocks = segment(grid)
            recipes = [
                _build_recipe(grid, block, i)
                for i, block in enumerate(blocks)
                if block.kind in (BlockKind.TABLE, BlockKind.CROSSTAB)
            ]
            sheets.append(
                SheetPlan(
                    sheet=worksheet.title,
                    fingerprint=sheet_fingerprint(blocks),
                    recipes=recipes,
                    blocks=blocks,
                    truncated_probe=grid.truncated,
                )
            )
    finally:
        workbook.close()

    if not any(sheet.recipes for sheet in sheets):
        raise IngestFailed(
            "No tabular blocks were detected in this workbook.",
            remedy="Check the sheet contains a header row followed by data rows.",
        )

    return FilePlan(path=path, filename=path.name, content_hash=content_hash(path), sheets=sheets)


def _locate_header(worksheet: Any, recipe: ExtractionRecipe) -> int:
    """Re-find the header by signature; absolute rows shift, signatures do not."""
    if not recipe.header_signature:
        return recipe.header_row
    wanted = {s.strip().lower() for s in recipe.header_signature}
    low = max(1, recipe.header_row + 1 - ANCHOR_SEARCH_SLACK)
    high = recipe.header_row + 1 + ANCHOR_SEARCH_SLACK
    for row in worksheet.iter_rows(
        min_row=low, max_row=high, min_col=recipe.first_col + 1, max_col=recipe.last_col + 1
    ):
        seen = {str(c.value).strip().lower() for c in row if c.value is not None}
        if wanted and len(wanted & seen) >= max(1, int(0.7 * len(wanted))):
            return row[0].row - 1
    return recipe.header_row


def _coerce_row(
    raw: tuple[Any, ...], specs: list[ColumnSpec], context: dict[str, Any]
) -> tuple[Any, ...]:
    out: list[Any] = []
    trace = False
    for index, spec in enumerate(specs):
        if spec.role is ColumnRole.DIMENSION and spec.source_name in context:
            supplied = context[spec.source_name]
            out.append(
                supplied.isoformat() if isinstance(supplied, (datetime, date)) else str(supplied)
            )
            continue
        value = raw[index] if index < len(raw) else None
        if spec.role is ColumnRole.MEASURE:
            parsed = parse_number(value)
            trace = trace or parsed.is_trace
            out.append(parsed.value if isinstance(parsed.value, float) else None)
        elif spec.role is ColumnRole.TIME:
            if isinstance(value, (datetime, date)):
                out.append(value.isoformat())
            else:
                parsed = parse_number(value).value
                # A year must not become 1991.0 -- it is a key, not a measurement.
                out.append(
                    int(parsed) if isinstance(parsed, float) and parsed.is_integer() else parsed
                )
        else:
            out.append(None if value is None else str(value).strip())
    out.append(trace)
    return tuple(out)


def stream(
    path: Path, recipe: ExtractionRecipe, batch_rows: int = BATCH_ROWS
) -> Iterator[RecordBatch]:
    """Replay one recipe over the full sheet in constant memory.

    Columns whose *header* is a period (Jan, 1901, a date) are melted into
    ``period`` / ``value`` pairs. This covers both a crosstab and the wide
    monthly rainfall series in one path: left wide, each new file would add
    columns instead of rows and the schema would never stabilise.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[recipe.sheet]
        header_row = _locate_header(worksheet, recipe)
        specs = recipe.stored_columns
        period_index = [i for i, s in enumerate(specs) if s.period_label]
        keep_index = [i for i, s in enumerate(specs) if not s.period_label]

        if period_index:
            columns = [specs[i].slug for i in keep_index] + [
                "period",
                "period_month",
                "value",
                "is_trace",
            ]
        else:
            columns = [s.slug for s in specs] + ["is_trace"]
        columns.append("_source_row")

        buffer: list[tuple[Any, ...]] = []
        excel_row = header_row + 1
        for row in worksheet.iter_rows(
            min_row=header_row + 2,
            min_col=recipe.first_col + 1,
            max_col=recipe.last_col + 1,
            values_only=True,
        ):
            excel_row += 1
            if all(v is None or str(v).strip() == "" for v in row):
                break  # the recipe extends until the first blank row
            first = row[0] if row else None
            if isinstance(first, str) and first.strip().lower().startswith("total"):
                continue  # retained separately as a checksum

            coerced = _coerce_row(row, specs, recipe.context)
            if period_index:
                keys = tuple(coerced[i] for i in keep_index)
                trace = bool(coerced[-1])
                for i in period_index:
                    label = specs[i].period_label or ""
                    buffer.append(
                        (*keys, label, MONTH_ABBR.get(label.lower()), coerced[i], trace, excel_row)
                    )
            else:
                buffer.append((*coerced, excel_row))

            if len(buffer) >= batch_rows:
                yield RecordBatch(recipe, columns, buffer)
                buffer = []
        if buffer:
            yield RecordBatch(recipe, columns, buffer)
    finally:
        workbook.close()
